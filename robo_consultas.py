#Busca Infos
import pandas as pd
import re
import openpyxl


#robotizacao
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
import os


def config_chrome():
    
    chrome_driver_path = r'chrome-win64\chrome.exe'
    os.environ["PATH"] += os.pathsep + chrome_driver_path

    # Configurar opções do ChromeDriver para executar em modo headless e desabilitar algumas funcionalidades de segurança
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--disable-web-security")
    chrome_options.add_argument("--disable-features=IsolateOrigins,site-per-process")
    chrome_options.add_argument("--disable-site-isolation-trials")    

    # Inicializar o WebDriver do Selenium com as opções configuradas
    driver = webdriver.Chrome(options=chrome_options)
    
    return driver


# Função para extrair número do processo a partir da página de detalhes do medicamento
def extract_process_number(page_source):
    process_number = None
    soup = BeautifulSoup(page_source, 'html.parser')
    process_tag = soup.find('td', class_='text-center col-sm-2 ng-binding')
    if process_tag:
        process_number = process_tag.text.strip()
    return process_number

# Função para extrair nome do princípio ativo e nome comercial a partir da página de detalhes do medicamento
def extract_active_principle_and_commercial_name(page_source):
    active_principle = None
    commercial_name = None
    soup = BeautifulSoup(page_source, 'html.parser')
    
    # Extrair nome do princípio ativo
    active_principle_tag = soup.find('td', class_='text-center col-sm-1 ng-binding', attrs={'ng-show': 'tipoProduto == 1'})
    if active_principle_tag:
        active_principle = active_principle_tag.text.strip()
    
    # Extrair nome comercial
    commercial_name_tag = soup.find('td', class_='col-xs-2 ng-binding', attrs={'ng-click': 'detail(produto)'})
    if commercial_name_tag:
        commercial_name = commercial_name_tag.text.strip()
    
    return active_principle, commercial_name

# Função para extrair a situação do medicamento (Válido ou Caduco/Cancelado)
def extract_medication_status(page_source):
    status = None
    soup = BeautifulSoup(page_source, 'html.parser')
    status_tag = soup.find('td', class_='text-center col-sm-1 ng-binding', string=re.compile(r'(Válido|Caduco/Cancelado)'))
    if status_tag:
        status = status_tag.text.strip()
    return status

# Função para verificar se o medicamento está registrado na ANVISA
def check_anvisa_registration(anvisa_number,driver):
    try:
        # Montar a URL da página de detalhes do medicamento na ANVISA
        process_url = f'https://consultas.anvisa.gov.br/#/medicamentos/q/?numeroRegistro={anvisa_number}'
        
        # Tentar acessar a URL
        driver.get(process_url)
        
        # Aguardar um curto período de tempo para garantir que a página seja carregada
        time.sleep(6)
        
        # Extrair nome do princípio ativo, nome comercial e situação do medicamento
        active_principle, commercial_name = extract_active_principle_and_commercial_name(driver.page_source)
        status = extract_medication_status(driver.page_source)
        
        # Extrair número do processo
        process_number = extract_process_number(driver.page_source)
        
        return process_number, active_principle, commercial_name, status
    except Exception as e:
        print(f"Erro ao verificar registro na ANVISA: {e}")
        return None, None, None, None

    
#retorna as informações encontradas através do google e do site da Anvisa
#recebe como entrada uma lista de tuplas (str_busca, dosagem)
#onde str_busca é o nome e dosagem tais e quais extraidos da sentença
def busca_medicamento(lista_medicamentos):
    # Inicializar o drive
    #driver = config_chrome()
    
    chrome_driver_path = r'chrome-win64\chrome.exe'
    os.environ["PATH"] += os.pathsep + chrome_driver_path

    # Configurar opções do ChromeDriver para executar em modo headless e desabilitar algumas funcionalidades de segurança
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--disable-web-security")
    chrome_options.add_argument("--disable-features=IsolateOrigins,site-per-process")
    chrome_options.add_argument("--disable-site-isolation-trials")    

    # Inicializar o WebDriver do Selenium com as opções configuradas
    driver = webdriver.Chrome(options=chrome_options)
    # Lista para armazenar os resultados
    results = list()
    # Loop através dos medicamentos
    for medication_name, dosagem in lista_medicamentos:
        # Verificar se a entrada não está vazia
        if medication_name.strip():  # Verifica se a string não está vazia após remoção dos espaços em branco
           
            # Pesquisar no Google apenas pelo primeiro resultado
            driver.get(f'https://www.google.com/search?q={medication_name}+numero+registro+ANVISA&num=1')
            
            # Aguardar até que os resultados da pesquisa estejam disponíveis
            time.sleep(6)
            
            # Extrair o HTML da página
            page_source = driver.page_source
            
            # Extrair o número de registro ANVISA
            soup = BeautifulSoup(page_source, 'html.parser')
            search_results = soup.find_all(class_='g')
            anvisa_number = None
            for result in search_results:
                if 'ANVISA' in result.text:
                    # Procurar por padrões de números de registro
                    matches = re.findall(r'(\d{9})', result.text)
                    if matches:
                        anvisa_number = matches[0]
                        break   
                        
            driver.get(f'https://www.google.com/search?q={medication_name}+principio+ativo+ANVISA&num=1')

            # Aguardar até que os resultados da pesquisa estejam disponíveis
            time.sleep(6)
            
            # Extrair o HTML da página
            page_source = driver.page_source
            
            # Busca no google o princípio ativo 
            soup = BeautifulSoup(page_source, 'html.parser')
            search_results = soup.find_all(class_='g')
            active_ingredient = None
            for result in search_results:
                if 'Princípio Ativo,' in result.text:
                    match = re.search(r'Ativo,(.*?), Medicamento', result.text)
                    if match:
                        active_ingredient = match.group(1).strip()
                        break
                    
            
            # Verificar se o número de registro é válido e se o medicamento está registrado na ANVISA
            if anvisa_number:
                process_number, active_principle, commercial_name, status = check_anvisa_registration(anvisa_number,driver)
                if process_number:
                    # Verificar se todos os campos necessários estão preenchidos
                    if active_principle or commercial_name:
                        registered_in_anvisa = "Sim"
                    else:
                        registered_in_anvisa = "Não"
                        active_principle = active_ingredient
                else:
                    registered_in_anvisa = "Não"
                    active_principle = active_ingredient
            else:
                process_number = None
                active_principle = active_ingredient
                commercial_name = None
                status = None
                registered_in_anvisa = "Não"
        else:
            # Entrada vazia, pular para o próximo medicamento
            continue    
    # Salvar o resultado
        #print(medication_name, commercial_name, active_principle,anvisa_number, process_number, registered_in_anvisa,status)
        results.append((medication_name, commercial_name, active_principle,anvisa_number, process_number, registered_in_anvisa,status))
    return results


#retorna as informações encontradas na tabela CMED
#recebe como entrada uma lista de tuplas (str_busca, nom_comerc, principio, num_reg, num_proc, x, y)
#onde str_busca é o nome tal qual extraido da sentença
def busca_CMED(m, lista_medicamentos):
    # Tabela de onde vão ser retiradas as informações dos medicamentos
    tabela_precos = pd.read_excel('tabela_CMED.xls', skiprows=52)
    
    # Lista que retorna no final da função
    infos_medicamentos = []

    # Laço para pegar as informações de cada medicamento na tabela pmvg
    for idx, (nom_origi, nom_comerc, medicamento, num_reg, num_proc, x, y) in enumerate(m):
        tabela_precos_2 = tabela_precos
        tabela_precos_2['SUBSTÂNCIA'] = tabela_precos_2['SUBSTÂNCIA'].fillna('')
        
        if medicamento is None:
            # Caso a lógica para buscar o princípio ativo não tenha funcionado, vamos procurar pela palavra trazida no GPT
            medicamento = lista_medicamentos[idx]
            medicamento = medicamento[0]
            #print('não funcionou para',medicamento)
        #else:
        #    print('funcionou para', medicamento)
        
        
        medicamento = medicamento.split(', ')
        for i in medicamento:
            # Compilando a expressão regular para correspondência exata de palavra
            padrao = re.compile(r'\b' + re.escape(i) + r'\b', re.IGNORECASE)
            tabela_precos_2 = tabela_precos_2[tabela_precos_2['SUBSTÂNCIA'].str.contains(padrao)]
        
        if tabela_precos_2.empty:
            # Pegar o primeiro elemento da lista medicamento
            primeiro_medicamento = medicamento[0] if isinstance(medicamento, list) else medicamento
            infos_medicamentos.append((primeiro_medicamento, nom_comerc, num_reg, 0))
        else:
            indice_maior_preco = tabela_precos_2['PMVG Sem Imposto'].idxmax()
            preco = tabela_precos_2.loc[indice_maior_preco, 'PMVG Sem Imposto']
            dosagem = tabela_precos_2.loc[indice_maior_preco, 'APRESENTAÇÃO']
            num_registro = tabela_precos_2.loc[indice_maior_preco, 'REGISTRO']
            nome_comercial = tabela_precos_2.loc[indice_maior_preco, 'PRODUTO'] 
            substancia = tabela_precos_2.loc[indice_maior_preco, 'SUBSTÂNCIA']
            #situacao especial em que não foi possível obter o preco
            if preco == None:
                preco = 0
            infos_medicamentos.append((substancia, nome_comercial, num_registro, preco))
        
    
    return infos_medicamentos

            
"""   
if __name__ == '__main__':
 
    
    medicamentos_buscados =  pd.read_csv('medicamentos4.csv')
    lista_medicamentos = [tuple(x) for x in medicamentos_buscados.to_numpy()]
    m = busca_medicamento(lista_medicamentos)
    n = busca_CMED(m,lista_medicamentos)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    #header = ['SUBSTÂNCIA', 'DOSAGEM', 'NOME_COMERCIAL','NUMERO_REGISTRO','PREÇO']
    header = ['SUBSTÂNCIA', 'NOME_COMERCIAL','NUMERO_REGISTRO','PREÇO']
    for col_index, header_title in enumerate(header, start=1):
        sheet.cell(row=1, column=col_index, value=header_title)

    for row_index, row in enumerate(n, start=2):
        for col_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)

    workbook.save('Informacoes_Medicamentos.xlsx')
    #output_csv_path = r'result_normaliza.csv'

"""